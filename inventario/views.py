import json
from datetime import timedelta
import openpyxl
from xhtml2pdf import pisa
from io import BytesIO
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.generic import DetailView, TemplateView
from django.views.decorators.http import require_POST
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Count, Q, Sum
from django.template.loader import render_to_string
from django.conf import settings
from dateutil.relativedelta import relativedelta
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import (
    Equipo, Cliente, Asignacion, CambioReparacion,
    HojaResponsabilidad, Accesorio, Alerta, Evidencia,
    Ubicacion, Categoria, SoftwareLicencia, MantenimientoPreventivo
)


# ========== 1. DASHBOARD MEJORADO ==========

class DashboardView(TemplateView):
    template_name = 'inventario/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        hoy = timezone.now().date()

        ctx['total_equipos'] = Equipo.objects.count()
        ctx['asignados'] = Equipo.objects.filter(estado='asignado').count()
        ctx['en_reparacion'] = Equipo.objects.filter(estado='en_reparacion').count()
        ctx['disponibles'] = Equipo.objects.filter(estado='disponible').count()
        ctx['dados_baja'] = Equipo.objects.filter(estado='dado_de_baja').count()
        ctx['clientes'] = Cliente.objects.count()
        ctx['asignaciones_activas'] = Asignacion.objects.filter(activa=True).count()
        ctx['hojas_pendientes'] = HojaResponsabilidad.objects.filter(firmado=False).count()
        ctx['alertas'] = Alerta.objects.filter(leida=False)[:8]
        ctx['total_alertas'] = Alerta.objects.filter(leida=False).count()
        ctx['stock_bajo'] = [a for a in Accesorio.objects.all() if a.stock_bajo()]
        ctx['ubicaciones'] = Ubicacion.objects.filter(activa=True).annotate(
            num_equipos=Count('equipos')
        ).order_by('-num_equipos')[:6]
        ctx['categorias'] = Categoria.objects.annotate(
            num_equipos=Count('equipos')
        ).order_by('-num_equipos')[:6]
        ctx['licencias_por_vencer'] = SoftwareLicencia.objects.filter(
            activa=True,
            fecha_vencimiento__lte=hoy + relativedelta(days=30),
            fecha_vencimiento__gte=hoy
        ).select_related('equipo')[:5]
        ctx['mantenimientos_proximos'] = MantenimientoPreventivo.objects.filter(
            completado=False,
            proxima_fecha__lte=hoy + relativedelta(days=7),
            proxima_fecha__gte=hoy
        ).select_related('equipo')[:5]
        ctx['equipos_fallas'] = Equipo.objects.annotate(
            num_fallas=Count('cambios')
        ).filter(num_fallas__gt=0).order_by('-num_fallas')[:5]
        ctx['clientes_top'] = Cliente.objects.annotate(
            num_equipos=Count('asignaciones', filter=Q(asignaciones__activa=True))
        ).filter(num_equipos__gt=0).order_by('-num_equipos')[:5]
        ctx['costo_total_reparaciones'] = CambioReparacion.objects.aggregate(
            total=Sum('costo')
        )['total'] or 0

        return ctx


# ========== 2. API STATS MEJORADA ==========

@api_view(['GET'])
def dashboard_stats_api(request):
    hoy = timezone.now()
    total = Equipo.objects.count()
    asignados = Equipo.objects.filter(estado='asignado').count()
    en_reparacion = Equipo.objects.filter(estado='en_reparacion').count()
    disponibles = Equipo.objects.filter(estado='disponible').count()
    dado_baja = Equipo.objects.filter(estado='dado_de_baja').count()

    por_marca = list(Equipo.objects.values('marca').annotate(total=Count('id')).order_by('-total')[:8])

    # --- CATEGORÍAS (con fallback) ---
    categorias_con_equipos = list(
        Categoria.objects
        .annotate(total=Count('equipos'))
        .filter(total__gt=0)
        .values('nombre', 'total')
        .order_by('-total')[:6]
    )
    sin_categoria = Equipo.objects.filter(categoria__isnull=True).count()
    if sin_categoria > 0:
        categorias_con_equipos.append({
            'nombre': '⚠️ Sin categoría',
            'total': sin_categoria
        })
    por_categoria = categorias_con_equipos

    # --- UBICACIONES (con fallback) ---
    ubicaciones_con_equipos = list(
        Ubicacion.objects
        .filter(activa=True)
        .annotate(total=Count('equipos'))
        .filter(total__gt=0)
        .values('nombre', 'total')
        .order_by('-total')[:6]
    )
    sin_ubicacion = Equipo.objects.filter(ubicacion__isnull=True).count()
    if sin_ubicacion > 0:
        ubicaciones_con_equipos.append({
            'nombre': '⚠️ Sin ubicación',
            'total': sin_ubicacion
        })
    por_ubicacion = ubicaciones_con_equipos

    asignaciones_mes = []
    for i in range(5, -1, -1):
        mes = hoy - relativedelta(months=i)
        count = Asignacion.objects.filter(fecha_asignacion__year=mes.year, fecha_asignacion__month=mes.month).count()
        asignaciones_mes.append({'mes': mes.strftime('%b %Y'), 'total': count})

    reparaciones_tipo = list(CambioReparacion.objects.values('tipo').annotate(total=Count('id')))

    licencias_por_tipo = list(SoftwareLicencia.objects.filter(activa=True).values('tipo').annotate(total=Count('id')))
    mantenimientos_estado = {
        'vencidos': MantenimientoPreventivo.objects.filter(completado=False, proxima_fecha__lt=hoy.date()).count(),
        'proximos_7d': MantenimientoPreventivo.objects.filter(completado=False, proxima_fecha__lte=hoy.date() + relativedelta(days=7), proxima_fecha__gte=hoy.date()).count(),
        'completados': MantenimientoPreventivo.objects.filter(completado=True).count(),
    }

    return Response({
        'totales': {'total': total, 'asignados': asignados, 'reparacion': en_reparacion, 'disponibles': disponibles, 'baja': dado_baja},
        'por_marca': por_marca,
        'por_categoria': por_categoria,
        'por_ubicacion': por_ubicacion,
        'asignaciones_trend': asignaciones_mes,
        'reparaciones_tipo': reparaciones_tipo,
        'licencias_por_tipo': licencias_por_tipo,
        'mantenimientos': mantenimientos_estado,
    })

# ========== 3. FICHA PUBLICA QR ==========

class EquipoFichaPublicaView(DetailView):
    model = Equipo
    template_name = 'inventario/equipo_ficha.html'
    context_object_name = 'equipo'
    slug_field = 'uuid'
    slug_url_kwarg = 'uuid'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['asignacion_actual'] = self.object.asignaciones.filter(activa=True).first()
        ctx['historial_reparaciones'] = self.object.cambios.all()[:10]
        ctx['historial_completo'] = self.object.history.all()[:10]
        ctx['evidencias'] = self.object.evidencias.all()[:6]
        ctx['licencias'] = self.object.licencias.filter(activa=True)
        ctx['mantenimientos'] = self.object.mantenimientos_preventivos.filter(completado=False)
        return ctx


# ========== 4. ESCANEAR QR ==========

def escanear_qr(request):
    return render(request, 'inventario/escanear_qr.html')


# ========== 5. BUSQUEDA GLOBAL ==========

@staff_member_required
def busqueda_global(request):
    q = request.GET.get('q', '').strip()
    resultados = {'equipos': [], 'clientes': [], 'asignaciones': [], 'reparaciones': []}
    if q:
        resultados['equipos'] = Equipo.objects.filter(
            Q(nombre__icontains=q) | Q(serial__icontains=q) | Q(marca__icontains=q) | Q(modelo__icontains=q)
        )[:10]
        resultados['clientes'] = Cliente.objects.filter(
            Q(nombre__icontains=q) | Q(dpi__icontains=q) | Q(email__icontains=q)
        )[:10]
        resultados['asignaciones'] = Asignacion.objects.filter(
            Q(equipo__nombre__icontains=q) | Q(equipo__serial__icontains=q) | Q(cliente__nombre__icontains=q)
        )[:10]
        resultados['reparaciones'] = CambioReparacion.objects.filter(
            Q(equipo__nombre__icontains=q) | Q(equipo__serial__icontains=q) | Q(descripcion__icontains=q)
        )[:10]
    return render(request, 'inventario/busqueda.html', {'q': q, 'resultados': resultados})


# ========== 6. REPORTES CON FILTROS ==========

@staff_member_required
def reporte_equipos(request):
    equipos = Equipo.objects.all()
    marca = request.GET.get('marca', '')
    estado = request.GET.get('estado', '')
    categoria = request.GET.get('categoria', '')
    ubicacion = request.GET.get('ubicacion', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if marca:
        equipos = equipos.filter(marca__icontains=marca)
    if estado:
        equipos = equipos.filter(estado=estado)
    if categoria:
        equipos = equipos.filter(categoria_id=categoria)
    if ubicacion:
        equipos = equipos.filter(ubicacion_id=ubicacion)
    if fecha_desde:
        equipos = equipos.filter(fecha_registro__date__gte=fecha_desde)
    if fecha_hasta:
        equipos = equipos.filter(fecha_registro__date__lte=fecha_hasta)

    if request.GET.get('exportar') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Equipos'
        ws.append(['Nombre', 'Categoria', 'Marca', 'Modelo', 'Serial', 'Ubicacion', 'Estado', 'Fecha Registro'])
        for e in equipos:
            ws.append([
                e.nombre,
                e.categoria.nombre if e.categoria else '—',
                e.marca, e.modelo, e.serial,
                e.ubicacion.nombre if e.ubicacion else '—',
                e.get_estado_display(),
                e.fecha_registro.strftime('%d/%m/%Y')
            ])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_equipos.xlsx'
        return response

    if request.GET.get('exportar') == 'pdf':
        html_string = render_to_string('reportes/reporte_equipos.html', {'equipos': equipos, 'filtros': request.GET})
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
        if pisa_status.err:
            return HttpResponse("Error generando PDF", status=500)
        pdf_buffer.seek(0)
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=reporte_equipos.pdf'
        return response

    marcas = Equipo.objects.exclude(marca='').values_list('marca', flat=True).distinct()
    categorias = Categoria.objects.all()
    ubicaciones = Ubicacion.objects.filter(activa=True)

    return render(request, 'inventario/reporte_equipos.html', {
        'equipos': equipos, 'marcas': marcas, 'categorias': categorias,
        'ubicaciones': ubicaciones, 'filtros': request.GET
    })


# ========== 7. EVIDENCIA ==========

@staff_member_required
def subir_evidencia(request):
    if request.method == 'POST':
        equipo_id = request.POST.get('equipo')
        tipo = request.POST.get('tipo')
        descripcion = request.POST.get('descripcion')
        imagen = request.FILES.get('imagen')
        if equipo_id and imagen:
            equipo = get_object_or_404(Equipo, pk=equipo_id)
            Evidencia.objects.create(
                equipo=equipo, tipo=tipo, descripcion=descripcion,
                imagen=imagen, subido_por=request.user
            )
            messages.success(request, 'Evidencia subida correctamente.')
        return redirect('subir_evidencia')
    equipos = Equipo.objects.all()
    return render(request, 'inventario/subir_evidencia.html', {'equipos': equipos})


# ========== 8. PDF HOJA RESPONSABILIDAD ==========

@staff_member_required
def generar_pdf_hoja(request, pk):
    hoja = get_object_or_404(HojaResponsabilidad, pk=pk)
    html_string = render_to_string('reportes/hoja_responsabilidad.html', {
        'hoja': hoja, 'asignacion': hoja.asignacion,
        'equipo': hoja.asignacion.equipo, 'cliente': hoja.asignacion.cliente,
        'fecha': timezone.now(),
    })
    pdf_buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
    if pisa_status.err:
        return HttpResponse("Error generando PDF", status=500)
    pdf_buffer.seek(0)
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=hoja_{hoja.asignacion.equipo.serial}.pdf'
    return response


# ========== 9. PAGINA FIRMA ==========

def pagina_firma(request, pk):
    hoja = get_object_or_404(HojaResponsabilidad, pk=pk)
    return render(request, 'inventario/firmar_hoja.html', {'hoja': hoja})


# ========== 10. FIRMA DIGITAL ==========

@require_POST
def firmar_hoja(request, pk):
    hoja = get_object_or_404(HojaResponsabilidad, pk=pk)
    import base64
    from django.core.files.base import ContentFile
    data = json.loads(request.body)
    firma_data = data.get('firma')
    if firma_data:
        fmt, imgstr = firma_data.split(';base64,')
        ext = fmt.split('/')[-1]
        hoja.firma_imagen.save(f'firma_{hoja.id}.{ext}', ContentFile(base64.b64decode(imgstr)), save=False)
        hoja.firmado = True
        hoja.fecha_firma = timezone.now()
        hoja.save()
        return JsonResponse({'ok': True})
    return JsonResponse({'ok': False, 'error': 'No se recibio firma'})


# ========== 11. IMPORTAR EXCEL ==========

@staff_member_required
def importar_equipos_excel(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'No se selecciono archivo')
            return redirect('admin:inventario_equipo_changelist')
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        creados = 0
        errores = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                nombre, marca, modelo, serial = row[0], row[1], row[2], row[3]
                if not serial:
                    continue
                Equipo.objects.create(nombre=nombre or 'Sin nombre', marca=marca or '', modelo=modelo or '', serial=str(serial))
                creados += 1
            except Exception as e:
                errores.append(f"Fila {idx}: {str(e)}")
        messages.success(request, f'{creados} equipos importados.')
        if errores:
            messages.warning(request, f'Errores: {", ".join(errores[:3])}')
        return redirect('admin:inventario_equipo_changelist')
    return render(request, 'admin/importar_equipos.html')


# ========== 12. ALERTAS MEJORADAS ==========

@staff_member_required
def verificar_alertas(request):
    hoy = timezone.now().date()
    creadas = 0

    for eq in Equipo.objects.filter(fecha_fin_garantia__lte=hoy+timedelta(days=30), fecha_fin_garantia__gte=hoy, estado__in=['disponible','asignado']):
        _, c = Alerta.objects.get_or_create(tipo='garantia', equipo=eq, defaults={'titulo':f'Garantia por vencer: {eq}','mensaje':f'La garantia del equipo {eq} vence el {eq.fecha_fin_garantia}.'})
        if c: creadas += 1

    for acc in Accesorio.objects.all():
        if acc.stock_bajo():
            _, c = Alerta.objects.get_or_create(tipo='stock', accesorio=acc, defaults={'titulo':f'Stock bajo: {acc.nombre}','mensaje':f'El accesorio {acc.nombre} tiene {acc.cantidad} unidades (minimo {acc.stock_minimo}).'})
            if c: creadas += 1

    hace_un_ano = hoy - timedelta(days=365)
    for asig in Asignacion.objects.filter(activa=True):
        fecha_asig = asig.fecha_asignacion.date() if hasattr(asig.fecha_asignacion, 'date') else asig.fecha_asignacion
        if fecha_asig <= hace_un_ano:
            if not asig.ultima_revision or asig.ultima_revision < hace_un_ano:
                _, c = Alerta.objects.get_or_create(tipo='revision', equipo=asig.equipo, defaults={'titulo':f'Revision pendiente: {asig.equipo}','mensaje':f'El equipo {asig.equipo} asignado a {asig.cliente} lleva mas de 1 ano sin revision.'})
                if c: creadas += 1

    for lic in SoftwareLicencia.objects.filter(activa=True, fecha_vencimiento__lte=hoy+timedelta(days=30), fecha_vencimiento__gte=hoy):
        _, c = Alerta.objects.get_or_create(tipo='licencia', licencia=lic, defaults={'titulo':f'Licencia por vencer: {lic.nombre}','mensaje':f'La licencia {lic.nombre} del equipo {lic.equipo} vence el {lic.fecha_vencimiento}.'})
        if c: creadas += 1

    for mp in MantenimientoPreventivo.objects.filter(completado=False, proxima_fecha__lte=hoy+timedelta(days=7)):
        _, c = Alerta.objects.get_or_create(tipo='mantenimiento', mantenimiento=mp, defaults={'titulo':f'Mantenimiento: {mp.titulo}','mensaje':f'El mantenimiento "{mp.titulo}" del equipo {mp.equipo} esta programado para el {mp.proxima_fecha}.'})
        if c: creadas += 1

    messages.success(request, f'Se verificaron alertas. {creadas} nuevas creadas.')
    return redirect('dashboard')


@require_POST
def marcar_alerta_leida(request, pk):
    alerta = get_object_or_404(Alerta, pk=pk)
    alerta.leida = True
    alerta.save()
    return JsonResponse({'ok': True})


# ========== 13. METRICAS AVANZADAS ==========

@staff_member_required
def metricas_avanzadas(request):
    reparaciones = CambioReparacion.objects.all()
    total_reparaciones = reparaciones.count()
    equipos_fallas = Equipo.objects.annotate(num_fallas=Count('cambios')).filter(num_fallas__gt=0).order_by('-num_fallas')[:5]
    clientes_top = Cliente.objects.annotate(num_equipos=Count('asignaciones', filter=Q(asignaciones__activa=True))).filter(num_equipos__gt=0).order_by('-num_equipos')[:5]
    costo_total = reparaciones.aggregate(total=Sum('costo'))['total'] or 0
    por_estado = {
        'disponible': Equipo.objects.filter(estado='disponible').count(),
        'asignado': Equipo.objects.filter(estado='asignado').count(),
        'en_reparacion': Equipo.objects.filter(estado='en_reparacion').count(),
        'dado_de_baja': Equipo.objects.filter(estado='dado_de_baja').count(),
    }
    costo_por_tipo = list(CambioReparacion.objects.values('tipo').annotate(total=Sum('costo')).order_by('-total'))
    equipos_por_categoria = list(Categoria.objects.annotate(total=Count('equipos')).values('nombre', 'total').order_by('-total'))
    equipos_por_ubicacion = list(Ubicacion.objects.filter(activa=True).annotate(total=Count('equipos')).values('nombre', 'total').order_by('-total'))
    licencias_activas = SoftwareLicencia.objects.filter(activa=True).count()
    licencias_por_vencer = SoftwareLicencia.objects.filter(activa=True, fecha_vencimiento__lte=timezone.now().date()+timedelta(days=30)).count()
    licencias_vencidas = SoftwareLicencia.objects.filter(activa=True, fecha_vencimiento__lt=timezone.now().date()).count()

    return render(request, 'inventario/metricas.html', {
        'total_reparaciones': total_reparaciones, 'equipos_fallas': equipos_fallas,
        'clientes_top': clientes_top, 'costo_total': costo_total, 'por_estado': por_estado,
        'costo_por_tipo': costo_por_tipo, 'equipos_por_categoria': equipos_por_categoria,
        'equipos_por_ubicacion': equipos_por_ubicacion, 'licencias_activas': licencias_activas,
        'licencias_por_vencer': licencias_por_vencer, 'licencias_vencidas': licencias_vencidas,
    })


# ========== 14. MANTENIMIENTOS PREVENTIVOS ==========

@staff_member_required
def lista_mantenimientos(request):
    hoy = timezone.now().date()
    proximos = MantenimientoPreventivo.objects.filter(completado=False, proxima_fecha__gte=hoy).select_related('equipo').order_by('proxima_fecha')
    vencidos = MantenimientoPreventivo.objects.filter(completado=False, proxima_fecha__lt=hoy).select_related('equipo').order_by('proxima_fecha')
    completados = MantenimientoPreventivo.objects.filter(completado=True).select_related('equipo').order_by('-ultima_fecha')[:20]
    return render(request, 'inventario/mantenimientos.html', {'proximos': proximos, 'vencidos': vencidos, 'completados': completados})


@staff_member_required
def completar_mantenimiento(request, pk):
    mp = get_object_or_404(MantenimientoPreventivo, pk=pk)
    mp.completado = True
    mp.ultima_fecha = timezone.now().date()
    mp.proxima_fecha = mp.calcular_proxima_fecha()
    mp.save()
    messages.success(request, f'Mantenimiento "{mp.titulo}" marcado como completado.')
    return redirect('lista_mantenimientos')


# ========== 15. LICENCIAS DE SOFTWARE ==========

@staff_member_required
def lista_licencias(request):
    hoy = timezone.now().date()
    activas = SoftwareLicencia.objects.filter(activa=True).select_related('equipo')
    por_vencer = activas.filter(fecha_vencimiento__lte=hoy+timedelta(days=30), fecha_vencimiento__gte=hoy)
    vencidas = activas.filter(fecha_vencimiento__lt=hoy)
    return render(request, 'inventario/licencias.html', {'activas': activas, 'por_vencer': por_vencer, 'vencidas': vencidas})


# ========== 16. EVIDENCIA MOVIL ==========

def subir_evidencia_movil(request):
    return render(request, 'inventario/subir_evidencia_movil.html')


# ========== 17. REPORTE EVIDENCIAS PDF ==========

import os
import requests


def reporte_evidencias_pdf(request):
    equipo_id = request.GET.get('equipo')
    if not equipo_id:
        return HttpResponse("Debes enviar el parametro ?equipo=ID o SERIAL", status=400)

    equipo = None
    try:
        equipo = Equipo.objects.get(serial=equipo_id)
    except Equipo.DoesNotExist:
        try:
            equipo_id_num = int(equipo_id)
            equipo = Equipo.objects.get(id=equipo_id_num)
        except (ValueError, Equipo.DoesNotExist):
            return HttpResponse(f"Equipo no encontrado: {equipo_id}", status=404)

    evidencias = Evidencia.objects.filter(equipo=equipo).order_by('-fecha')

    qr_base64 = None
    try:
        import qrcode
        import base64
        site_url = getattr(settings, 'SITE_URL', 'https://inventario-equipos-hkmd.onrender.com')
        reporte_url = f"{site_url}/api/evidencias/reporte/?equipo={equipo.serial}"
        qr = qrcode.make(reporte_url, box_size=6, border=2)
        buffer = BytesIO()
        qr.save(buffer, format='PNG')
        qr_base64 = f"data:image/png;base64,{base64.b64encode(buffer.getvalue()).decode()}"
    except Exception:
        qr_base64 = None

    html_string = render_to_string('inventario/reporte_evidencias_pdf.html', {
        'equipo': equipo, 'evidencias': evidencias,
        'total_evidencias': evidencias.count(),
        'fecha_generacion': timezone.now(), 'qr_base64': qr_base64,
    })

    pdf_buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
    if pisa_status.err:
        return HttpResponse("Error generando PDF", status=500)
    pdf_buffer.seek(0)

    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="reporte_evidencias_{equipo.serial}.pdf"'
    return response


def seed_licencias_mantenimientos_web(request):
    """Endpoint temporal para ejecutar seed en Render (plan gratuito sin shell)."""
    TOKEN_SECRETO = "renderseed2024"
    token = request.GET.get("token", "")
    
    if token != TOKEN_SECRETO:
        return HttpResponse("Acceso no autorizado", status=403)
    
    from datetime import date, timedelta
    from decimal import Decimal
    import random
    from inventario.models import Equipo, SoftwareLicencia, MantenimientoPreventivo, Alerta
    
    equipos = list(Equipo.objects.all())
    if not equipos:
        return HttpResponse("<h1>❌ No hay equipos en la base de datos</h1>", status=400)
    
    resultados = []
    resultados.append(f"<p>Procesando {len(equipos)} equipos...</p>")
    
    # LICENCIAS
    licencias_data = [
        ("Windows 11 Pro", "os", 365),
        ("Microsoft Office 365", "office", 365),
        ("Kaspersky Endpoint Security", "antivirus", 365),
        ("AutoCAD 2024", "cad", 365),
        ("Adobe Creative Cloud", "otro", 365),
        ("SQL Server 2022", "db", 730),
        ("VMware Workstation", "otro", 730),
        ("Norton Antivirus", "antivirus", 365),
        ("Windows Server 2022", "os", 730),
        ("Microsoft Project", "office", 365),
        ("Autodesk Revit", "cad", 365),
        ("Bitdefender GravityZone", "antivirus", 365),
    ]
    
    licencias_creadas = 0
    for equipo in equipos:
        for _ in range(random.randint(1, 3)):
            nombre, tipo, dias = random.choice(licencias_data)
            if random.random() < 0.3:
                inicio = date.today() - timedelta(days=random.randint(300, 340))
            else:
                inicio = date.today() - timedelta(days=random.randint(0, 300))
            venc = inicio + timedelta(days=dias)
            clave = f"XXXXX-XXXXX-XXXXX-XXXXX-{random.randint(10000,99999)}"
            costo = Decimal(random.randint(500, 8000))
            
            obj, created = SoftwareLicencia.objects.get_or_create(
                equipo=equipo, nombre=nombre,
                defaults={"tipo": tipo, "clave": clave, "fecha_inicio": inicio,
                          "fecha_vencimiento": venc, "costo": costo, "activa": True}
            )
            if created:
                licencias_creadas += 1
    
    resultados.append(f"<p>✅ <strong>{licencias_creadas}</strong> licencias de software creadas</p>")
    
    # MANTENIMIENTOS
    mantenimientos_data = [
        ("Limpieza interna y cambio de pasta termica", "trimestral"),
        ("Revision de bateria y calibracion", "semestral"),
        ("Actualizacion de firmware y drivers", "trimestral"),
        ("Limpieza de ventiladores y disipadores", "mensual"),
        ("Revision de cables y conectores", "semestral"),
        ("Escaneo de virus y optimizacion", "mensual"),
        ("Respaldo de datos y verificacion", "semanal"),
        ("Revision de rendimiento y temperatura", "mensual"),
        ("Limpieza de pantalla y teclado", "semanal"),
        ("Revision de garantia y estado fisico", "anual"),
    ]
    tecnicos = ["Luis Torres", "Sandra Morales", "Fernando Paz", "Carlos Mendez",
                "Ana Lopez", "Diana Herrera", "Pedro Ruiz"]
    
    mant_creados = 0
    for equipo in equipos:
        for _ in range(random.randint(1, 2)):
            titulo, freq = random.choice(mantenimientos_data)
            tecnico = random.choice(tecnicos)
            completado = random.random() < 0.4
            if completado:
                ultima = date.today() - timedelta(days=random.randint(1, 60))
            else:
                ultima = date.today() - timedelta(days=random.randint(-30, 90))
            
            obj, created = MantenimientoPreventivo.objects.get_or_create(
                equipo=equipo, titulo=titulo,
                defaults={"descripcion": f"{titulo} para {equipo.nombre}.",
                          "frecuencia": freq,
                          "ultima_fecha": ultima if completado else (ultima if ultima < date.today() else None),
                          "tecnico": tecnico, "completado": completado}
            )
            if created:
                mant_creados += 1
    
    resultados.append(f"<p>✅ <strong>{mant_creados}</strong> mantenimientos preventivos creados</p>")
    
    # ALERTAS
    hoy = date.today()
    alertas = 0
    for lic in SoftwareLicencia.objects.filter(activa=True, fecha_vencimiento__lte=hoy+timedelta(days=30), fecha_vencimiento__gte=hoy):
        _, c = Alerta.objects.get_or_create(
            tipo="licencia", licencia=lic,
            defaults={"titulo": f"Licencia por vencer: {lic.nombre}",
                      "mensaje": f"La licencia {lic.nombre} de {lic.equipo} vence el {lic.fecha_vencimiento}."})
        if c:
            alertas += 1
    
    for mp in MantenimientoPreventivo.objects.filter(completado=False, proxima_fecha__lte=hoy+timedelta(days=7)):
        _, c = Alerta.objects.get_or_create(
            tipo="mantenimiento", mantenimiento=mp,
            defaults={"titulo": f"Mantenimiento: {mp.titulo}",
                      "mensaje": f'El mantenimiento "{mp.titulo}" de {mp.equipo} esta programado para el {mp.proxima_fecha}.'})
        if c:
            alertas += 1
    
    resultados.append(f"<p>✅ <strong>{alertas}</strong> alertas generadas</p>")
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head><title>Seed Completado</title>
    <style>
        body {{ font-family: Arial, sans-serif; max-width: 600px; margin: 50px auto; padding: 20px; }}
        .ok {{ color: green; }}
        h1 {{ color: #2c3e50; }}
        .box {{ background: #f8f9fa; padding: 20px; border-radius: 8px; border-left: 4px solid #198754; }}
    </style>
    </head>
    <body>
        <h1>🌱 Seed Ejecutado en Render</h1>
        <div class="box">
            {''.join(resultados)}
        </div>
        <p style="margin-top: 20px;">
            <a href="/" style="padding: 10px 20px; background: #0d6efd; color: white; text-decoration: none; border-radius: 5px;">Ir al Dashboard</a>
            <a href="/metricas/" style="padding: 10px 20px; background: #6c757d; color: white; text-decoration: none; border-radius: 5px; margin-left: 10px;">Ver Metricas</a>
        </p>
        <p style="color: #dc3545; font-size: 12px; margin-top: 30px;">
            ⚠️ <strong>Importante:</strong> Elimina este endpoint despues de usarlo para evitar que se ejecute multiples veces.
        </p>
    </body>
    </html>
    """
    return HttpResponse(html)


# ========== 18. MANUAL DE USUARIO ==========

def manual_usuario(request):
    """Renderiza el Manual de Usuario del sistema."""
    return render(request, 'inventario/manual.html')
